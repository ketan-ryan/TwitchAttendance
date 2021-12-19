from openpyxl.styles import PatternFill
from twitch_irc import TwitchIRC
import datetime
import openpyxl
import os


class SheetManager:
    wb = None
    current_sheet = None
    need_list = False

    day = datetime.datetime.today().strftime('%-d')
    name = None
    total_entries = 2
    all_followers = []
    mods = []
    vips = []

    green = PatternFill(fgColor='00ff00', fill_type='solid')    # present
    pink = PatternFill(fgColor='ff00e6', fill_type='solid')     # vip
    orange = PatternFill(fgColor='ff5100', fill_type='solid')   # lurking
    blue = PatternFill(fgColor='0000ff', fill_type='solid')     # moderator


    def __init__(self, streamer_name) -> None:
        """
        Create Excel book if none exists
        Create worksheet for current month-year
        Get list of mods and vips
        """
        self.name = streamer_name
        self.check_create()

        self.wb = openpyxl.load_workbook(f'{streamer_name}.xlsx')

        self.current_sheet = self.wb.active
        month_year = datetime.datetime.today().strftime('%m-%Y')
        sheet_exists = False

        for idx, s in enumerate(self.wb.sheetnames):
            if s == month_year:
                self.current_sheet = self.wb.worksheets[idx]
                sheet_exists = True

        if not sheet_exists:
            self.wb.active = self.wb.create_sheet(month_year)
            self.current_sheet = self.wb.active

            self.current_sheet['A1'] = 'Chatter'
            self.need_list = True

        self.current_sheet.cell(1, int(self.day) + 1).value = datetime.datetime.today().strftime('%-d/%-m')
        irc = TwitchIRC(self.name)
        self.mods = irc.get_mods()
        self.vips = irc.get_vips()
        irc.close()


    def check_create(self):
        """
        Check if an Excel sheet exists already for the streamer,
        Create one if not
        """
        if not os.path.isfile(f'{self.name}.xlsx'):
            self.wb = openpyxl.Workbook()
            self.wb.save(f'{self.name}.xlsx')


    def need_update(self):
        """Getter method"""
        return self.need_list


    def write_list(self, chatter_names):
        """
        Fill column A with all followers of the Twitch stream
        If a follower is a moderator, highlight them blue
        If a follower is a vip, highlight them pink
        """
        for idx, name in enumerate(chatter_names):
            cell = self.current_sheet.cell(self.total_entries + idx, column=1)
            cell.value = name

            if name.lower() in self.mods:
                cell.fill = self.blue
            elif name.lower() in self.vips:
                cell.fill = self.pink

        self.total_entries = self.total_entries + len(chatter_names)
        print(self.total_entries)


    def get_followers(self):
        """Gets list of all followers from excel sheet in lowercase for comparison checking"""
        if(len(self.all_followers) == 0):
            for i in range(2, self.current_sheet.max_row):
                self.all_followers.append(str(self.current_sheet.cell(i, 1).value))
        self.all_followers = [str.lower(follower) for follower in self.all_followers]
        return self.all_followers


    def update_attendance(self, username, val):
        """
        Updates attendance for current chatter
        If lurking, only update if chatter doesn't already have an entry (lurking has lower priority)
        If present, overwrite current value with present
        Highlight lurkers orange
        Highlight present chatters green
        """
        try:
            idx = self.all_followers.index(username)
        except ValueError:
            return

        col = int(self.day) + 1
        cell = self.current_sheet.cell(idx + 2, col)
        if val == 'Lurking' and self.current_sheet.cell(idx + 2, col).value is None:
            cell.value = val
            cell.fill = self.orange
        elif val == 'Present':
            cell.value = val
            cell.fill = self.green


    def close(self):
        """Save workbook"""
        self.wb.save(f'{self.name}.xlsx')
